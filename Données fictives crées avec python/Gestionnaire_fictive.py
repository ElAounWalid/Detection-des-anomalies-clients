import pandas as pd
import openpyxl
import random


noms = ["Gharbi", "Ben Ali", "Toumi", "Larbi", "Hammami", "Mansour", "Haddad", "Ammar", "Gabsi", "Bouhlel", "Mabrouk", "Cherif", "Abdeljelil", "Chakroun", "Khemiri", "Boussaid", "Bouhamed", "Ben Ammar", "Zribi", "Abid", "Sahli", "Soltani", "Majri", "Othmani", "Mansouri", "Khalfallah", "Hajji", "Cherni", "Berriche", "Boukadida"]

prenoms = ["Ahmed", "Mohamed", "Ali", "Hassan", "Omar", "Khaled", "Mounir", "Kais", "Nizar", "Zied", "Nabil", "Majdi", "Samir", "Riadh", "Khalil", "Sami", "Anis", "Saber", "Adel", "Mehdi", "Wissem", "Mourad", "Mokhtar", "Abdelhamid", "Sofiene", "Sami", "Ameni", "Yosra", "Rahma", "Mariem","Marwa","Maram","Nour","Thouraya","Lilia","Ons","Maissa","Yassmine","Rania","Hanen"]

def generer_nom_prenom():
    nom = random.choice(noms)
    prenom = random.choice(prenoms)
    return [nom,prenom]

# Charger le fichier Excel
df = pd.read_excel('CLIENT.xlsx')

# Supprimer les doublons et mettre les valeurs dans une liste
liste_valeurs = df['CODE_GESTIONNAIRE'].unique().tolist()

wb = openpyxl.load_workbook('GESTIONNAIRE.xlsx')

# Sélection de la feuille de calcul
sheet = wb['Gestionnaire_F']

# Remplissage des cellules

for row_num in range(0,len(liste_valeurs)) :
    sheet.cell(row=row_num+2, column=1).value = liste_valeurs[row_num]
    sheet.cell(row=row_num+2, column=2).value = generer_nom_prenom()[0]
    sheet.cell(row=row_num+2, column=3).value = generer_nom_prenom()[1]
    
# Sauvegarde du fichier
wb.save('GESTIONNAIRE.xlsx')