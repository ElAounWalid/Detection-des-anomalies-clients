import openpyxl
import random


def generate_phone_number():
    """Génère un numéro de téléphone aléatoire"""
    phone_number = random.choice(["+216","216",""])
    for i in range(8):
        phone_number += str(random.randint(0, 9))  # Ajoute un chiffre aléatoire entre 0 et 9
        phone_number += random.choice([" ",""])  # Ajoute ou non un espace
    return phone_number


#ouverture de fichier Excel 
wb = openpyxl.load_workbook('TEL.xlsx')

# Sélection de la feuille de calcul
sheet = wb['Telephone_F']

# Remplissage des cellules
for row_num in range(2,2002):
    sheet.cell(row=row_num, column=11).value = generate_phone_number()
    sheet.cell(row=row_num, column=12).value = generate_phone_number()

# Sauvegarde du fichier
wb.save('TEL.xlsx')

