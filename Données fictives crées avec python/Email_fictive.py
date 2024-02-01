import openpyxl
import random
import string

def generate_Email(Nom,Prenom):
    """Génère un numéro de téléphone aléatoire"""
    identifiant = ''.join(random.choice(string.ascii_lowercase + string.digits) for _ in range(6))
    Email = str(Nom) + "." + str(Prenom) + identifiant + random.choice(["@gmail.com","@yahoo.fr","@hotmail.fr","@hotmail.com","@outlook.com"])
    return Email


#ouverture de fichier Excel 
wb = openpyxl.load_workbook('EMAIL.xlsx')
NomPrenom = openpyxl.load_workbook('NOM_PRENOM.xlsx')

# Sélection de la feuille de calcul
sheet = wb['Email_F']
sheet_NP = NomPrenom['Nom_F']

# Remplissage des cellules
for row_num in range(2,2002):
    nom = sheet_NP.cell(row=row_num, column=11).value
    prenom = sheet_NP.cell(row=row_num, column=12).value
    sheet.cell(row=row_num, column=11).value = generate_Email(nom,prenom)

# Sauvegarde du fichier
wb.save('EMAIL.xlsx')