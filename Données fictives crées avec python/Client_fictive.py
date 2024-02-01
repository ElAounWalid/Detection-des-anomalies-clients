import openpyxl
import random
from datetime import datetime, timedelta
import pandas as pd


#ouverture de fichier Excel 

wb = openpyxl.load_workbook('CLIENT.xlsx')

# Sélection de la feuille de calcul
sheet = wb['Client_F']

#Chrger le fichier AGENCE.xlsx qui contient la liste des agences avec leurs codes
df = pd.read_excel('AGENCE.xlsx')

#Selectionner la colonne qui contient les codes des agences
colonne = df['ID_AGENCE']


# Remplissage des cellules
for row_num in range(2,2002):
    sheet.cell(row=row_num, column=1).value = row_num-1
    sheet.cell(row=row_num, column=2).value = random.choice(["Particuliers","TRE","Profession liberale","TPE","Entreprise"])
    
    marche = sheet.cell(row=row_num, column=2).value
    if marche == "Particuliers":
        sheet.cell(row=row_num, column=3).value = random.choice(["P1","P2","P3","P4","P5"])
    elif marche == "TRE":
        sheet.cell(row=row_num, column=3).value = random.choice(["P6","P7","P8"])
    elif marche == "Profession liberale":
        sheet.cell(row=row_num, column=3).value = random.choice(["P9","P10","P11"])
    elif marche == "TPE":
        sheet.cell(row=row_num, column=3).value = random.choice(["P12","P13","P14","P15"])
    else :
        sheet.cell(row=row_num, column=3).value = random.choice(["P16","P17"])
        
        
    sheet.cell(row=row_num, column=4).value = random.choice(["S1","S2","S3","S4","S5"])
    sheet.cell(row=row_num, column=5).value = random.choice(["I1","I2","I3","I4","I5"])   
    sheet.cell(row=row_num, column=6).value = random.choice(["VIP","TR GEI","DEGITAL TSF","OFF SHORE","BANQUE PRIVEE","BANQUE DIGITALE"]) 

    # Définir la date de début
    start_date = datetime(2008, 1, 1)
    # Obtenir la date système
    end_date = datetime.now()

    #obtenir une date d'ouverture aleatoire
    # Calculer la différence entre les deux dates
    diff_dateDefaut_dateSys = end_date - start_date
    # Obtenir un nombre aléatoire entre 0 et le nombre de jours dans la différence de dates
    random_day = random.randint(0, diff_dateDefaut_dateSys.days)
    # Ajouter le nombre de jours aléatoire à la date de début pour obtenir la date aléatoire
    date_ouverture = start_date + timedelta(days=random_day)
    
    date_modification = date_ouverture + timedelta(days=40)
    
    sheet.cell(row=row_num, column=7).value = date_ouverture
    sheet.cell(row=row_num, column=8).value = date_modification
    sheet.cell(row=row_num, column=9).value = random.randint(100,999)
    sheet.cell(row=row_num, column=10).value = colonne.sample().values[0]
    
    
# Sauvegarde du fichier
wb.save('CLIENT.xlsx')