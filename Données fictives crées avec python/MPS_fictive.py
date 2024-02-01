import openpyxl

# Charger le fichier Excel
wb = openpyxl.load_workbook('CLIENT.xlsx')

# Sélection de la feuille de calcul
sheet = wb['Client_F']
# Creation d'une colonne qui contient le marche,la profession,le segment

print(sheet.max_row)

Segmentation_Liste = []
for row_num in range(2,2002) :
    Segmentation_chaine = str(sheet.cell(row=row_num, column=2).value) +","+str(sheet.cell(row=row_num, column=3).value)+","+str(sheet.cell(row=row_num, column=4).value)+","+str(sheet.cell(row=row_num, column=5).value)+","+str(sheet.cell(row=row_num, column=6).value)
    if Segmentation_chaine not in Segmentation_Liste:
        Segmentation_Liste.append(Segmentation_chaine)
        
wb = openpyxl.load_workbook('Segmentation.xlsx')

# Sélection de la feuille de calcul
sheet = wb['Segmentation_F']

# Remplissage des cellules

for row_num in range(0,len(Segmentation_Liste)) :
    sheet.cell(row=row_num+2, column=1).value = row_num+1
    sheet.cell(row=row_num+2, column=2).value = Segmentation_Liste[row_num].split(",")[0]
    sheet.cell(row=row_num+2, column=3).value = Segmentation_Liste[row_num].split(",")[1]
    sheet.cell(row=row_num+2, column=4).value = Segmentation_Liste[row_num].split(",")[2]
    sheet.cell(row=row_num+2, column=5).value = Segmentation_Liste[row_num].split(",")[3]
    sheet.cell(row=row_num+2, column=6).value = Segmentation_Liste[row_num].split(",")[4]
    
# Sauvegarde du fichier
wb.save('Segmentation.xlsx')