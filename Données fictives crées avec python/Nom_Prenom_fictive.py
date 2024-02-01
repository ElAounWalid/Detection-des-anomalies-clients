import openpyxl
import random

noms = ["Gharbi", "Ben Ali", "Toumi", "Larbi", "Hammami", "Mansour", "Haddad", "Ammar", "Gabsi", "Bouhlel", "Mabrouk", "Cherif", "Abdeljelil", "Chakroun", "Khemiri", "Boussaid", "Bouhamed", "Ben Ammar", "Zribi", "Abid", "Sahli", "Soltani", "Majri", "Othmani", "Mansouri", "Khalfallah", "Hajji", "Cherni", "Berriche", "Boukadida"]

prenoms = ["Ahmed", "Mohamed", "Ali", "Hassan", "Omar", "Khaled", "Mounir", "Kais", "Nizar", "Zied", "Nabil", "Majdi", "Samir", "Riadh", "Khalil", "Sami", "Anis", "Saber", "Adel", "Mehdi", "Wissem", "Mourad", "Mokhtar", "Abdelhamid", "Sofiene", "Sami", "Ameni", "Yosra", "Rahma", "Mariem","Marwa","Maram","Nour","Thouraya","Lilia","Ons","Maissa","Yassmine","Rania","Hanen"]

def generer_nom_prenom():
    nom = random.choice(noms)
    prenom = random.choice(prenoms)
    return [nom,prenom]

#ouverture de fichier Excel 
wb = openpyxl.load_workbook('NOM_PRENOM.xlsx')

# Sélection de la feuille de calcul
sheet = wb['Nom_F']

# Remplissage des cellules
for row_num in range(2,2002):
    sheet.cell(row=row_num, column=11).value = random.choice(["",generer_nom_prenom()[0]])
    sheet.cell(row=row_num, column=12).value = random.choice(["",generer_nom_prenom()[1]])

# Sauvegarde du fichier
wb.save('NOM_PRENOM.xlsx')
